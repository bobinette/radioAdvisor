#!usr/bin/python
# -*- coding: utf-8 -*-

# --------------------------------------------------------
# Copyright (c) 2018 RadioAdvisor
# Written by Yann Giret
# --------------------------------------------------------

import codecs
import csv
import os

from openpyxl import load_workbook


def get_csv_data():

    # Parse csv
    train_data = parse_csv("train")
    val_data = parse_csv("validation")
    # Store in one dict
    data = dict()
    data.update(train_data)
    data.update(val_data)

    return data


def parse_csv(set_type="train"):

    with open(os.path.join("menisque_%s_set.csv" % set_type), "rU") as f:
        raw_info = csv.reader(f, delimiter=',', dialect=csv.excel_tab)

        info_dict = {}
        cmpt = 0
        for row in raw_info:
            if cmpt == 0:
                cmpt += 1
                continue
            info_dict[row[0]] = {"corne_anterieure": bool(int(row[1])),
                                 "corne_posterieure": bool(int(row[2])),
                                 "orient_anterieure": str(row[3]),
                                 "orient_posterieure": str(row[4]),
                                 "position": str(row[5])}

    return info_dict


def load_seg_annotation(filename):

    wb = load_workbook("%s.xlsx" % filename)
    ws = wb.active
    db = {}
    for idx, row in enumerate(ws):
        if idx == 0:
            continue
        if row[0].value is None:
            break
        id_ = row[0].value.split(".")[0]
        db[id_] = {"data_file": row[0].value,
                   "annot_file": row[1].value,
                   "area": float(row[2].value),
                   "quality": row[3].value}

    return db


def convert_csv(filename):
    """Create the csv file based on the args.

    Each arg should be a list of size n, n being the number
    of images. Then:
    - f_scores[i]: a probability indicating broken or not
    - location[i]: 0 for antérieure, 1 for postérieure
    - o_scores[i]: an array of size 2: [<h prob>, <v prob>]
    """
    wb = load_workbook("%s.xlsx" % filename)
    ws = wb.active

    res = []
    for row in ws:
        # ant, post = (1, 0) if location == 0 else (0, 1)
        res.append(u'%s,%s,%s,%s,%s,%s' % (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))

    csv = u'\n'.join(res)
    with codecs.open('%s.csv' % filename, 'w', 'utf-8') as file:
        file.write('\ufeff')  # UTF-8 BOM header
        file.write(csv)
